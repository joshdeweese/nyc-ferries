from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('data/2020/PrivateFerryRidership_2020_12.xlsx')
print(wb.sheetnames)

ws_nyc = wb['NYC Ferry']

l = []
for row in range(1, ws_nyc.max_row + 1):
    r = [cell.value for cell in ws_nyc[row]]
    l.append(r)

df = pd.DataFrame(l)
df.to_clipboard()
